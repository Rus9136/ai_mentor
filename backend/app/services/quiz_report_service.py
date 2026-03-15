"""
Quiz Report service (Phase 2.3) — XLSX generation using openpyxl.
"""
import logging
from io import BytesIO
from typing import Optional

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.quiz import QuizSession, QuizParticipant, QuizAnswer, QuizSessionStatus
from app.models.student import Student
from app.models.user import User
from app.models.test import Test, Question
from app.repositories.quiz_repo import QuizRepository

logger = logging.getLogger(__name__)


class QuizReportService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = QuizRepository(db)

    async def class_report(self, session_id: int, teacher_id: int) -> BytesIO:
        """Generate class report: rows=students, columns=Score, Rank, Q1..QN, Correct%, AvgTime."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        session = await self.repo.get_session(session_id)
        if not session:
            raise ValueError("Session not found")
        if session.teacher_id != teacher_id:
            raise ValueError("Not the session owner")

        participants = await self.repo.get_participants_with_names(session_id)
        answers = await self.repo.get_answers_matrix(session_id)

        # Index answers
        answer_map: dict[int, dict[int, object]] = {}
        for a in answers:
            answer_map.setdefault(a.participant_id, {})[a.question_index] = a

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Class Report"

        # Header
        headers = ["#", "Ученик", "Очки", "Место"]
        for qi in range(session.question_count):
            headers.append(f"Q{qi + 1}")
        headers.extend(["Правильных", "%", "Ср. время (с)"])

        header_font = Font(bold=True)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font

        # Data rows
        for row_idx, data in enumerate(participants, 2):
            p = data["participant"]
            p_answers = answer_map.get(p.id, {})

            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=data["student_name"])
            ws.cell(row=row_idx, column=3, value=p.total_score)
            ws.cell(row=row_idx, column=4, value=p.rank or "—")

            total_time = 0
            answered_count = 0
            for qi in range(session.question_count):
                col = 5 + qi
                ans = p_answers.get(qi)
                if ans:
                    cell = ws.cell(row=row_idx, column=col, value="✓" if ans.is_correct else "✗")
                    cell.fill = green_fill if ans.is_correct else red_fill
                    total_time += ans.answer_time_ms
                    answered_count += 1
                else:
                    ws.cell(row=row_idx, column=col, value="—")

            correct_col = 5 + session.question_count
            ws.cell(row=row_idx, column=correct_col, value=p.correct_answers)
            pct = round(p.correct_answers / session.question_count * 100) if session.question_count > 0 else 0
            ws.cell(row=row_idx, column=correct_col + 1, value=f"{pct}%")
            avg_time = round(total_time / answered_count / 1000, 1) if answered_count > 0 else 0
            ws.cell(row=row_idx, column=correct_col + 2, value=avg_time)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def question_report(self, session_id: int, teacher_id: int) -> BytesIO:
        """Generate question analysis: rows=questions, columns=Difficulty, AvgTime, Distribution."""
        import openpyxl
        from openpyxl.styles import Font

        session = await self.repo.get_session(session_id)
        if not session:
            raise ValueError("Session not found")
        if session.teacher_id != teacher_id:
            raise ValueError("Not the session owner")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Question Analysis"

        headers = ["Вопрос", "Всего ответов", "Правильных", "% правильных", "Ср. время (с)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        for qi in range(session.question_count):
            stats = await self.repo.get_answer_stats(session_id, qi)
            # Count total and correct
            result = await self.db.execute(
                select(
                    func.count().label("total"),
                    func.sum(QuizAnswer.is_correct.cast(int)).label("correct"),
                    func.avg(QuizAnswer.answer_time_ms).label("avg_time"),
                ).where(
                    QuizAnswer.quiz_session_id == session_id,
                    QuizAnswer.question_index == qi,
                )
            )
            row = result.one()
            total = row[0] or 0
            correct = row[1] or 0
            avg_time = round((row[2] or 0) / 1000, 1)
            pct = round(correct / total * 100) if total > 0 else 0

            r = qi + 2
            ws.cell(row=r, column=1, value=f"Q{qi + 1}")
            ws.cell(row=r, column=2, value=total)
            ws.cell(row=r, column=3, value=correct)
            ws.cell(row=r, column=4, value=f"{pct}%")
            ws.cell(row=r, column=5, value=avg_time)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    async def trend_report(self, class_id: int, teacher_id: int) -> BytesIO:
        """Generate trend report: rows=sessions over time, columns=Date, AvgScore, Participation."""
        import openpyxl
        from openpyxl.styles import Font

        # Get recent finished sessions for this class
        result = await self.db.execute(
            select(QuizSession)
            .where(
                QuizSession.class_id == class_id,
                QuizSession.teacher_id == teacher_id,
                QuizSession.status == QuizSessionStatus.FINISHED,
            )
            .order_by(QuizSession.finished_at.desc())
            .limit(20)
        )
        sessions = list(result.scalars().all())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trend Report"

        headers = ["Дата", "Участников", "Вопросов", "Ср. очки", "Ср. правильных %"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        for row_idx, s in enumerate(reversed(sessions), 2):
            # Get participant stats
            participants = await self.repo.get_participants_with_names(s.id)
            count = len(participants)
            avg_score = sum(p["participant"].total_score for p in participants) / count if count > 0 else 0
            avg_correct_pct = sum(
                p["participant"].correct_answers / s.question_count * 100 if s.question_count > 0 else 0
                for p in participants
            ) / count if count > 0 else 0

            ws.cell(row=row_idx, column=1, value=s.finished_at.strftime("%Y-%m-%d %H:%M") if s.finished_at else "—")
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(row=row_idx, column=3, value=s.question_count)
            ws.cell(row=row_idx, column=4, value=round(avg_score))
            ws.cell(row=row_idx, column=5, value=f"{round(avg_correct_pct)}%")

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
